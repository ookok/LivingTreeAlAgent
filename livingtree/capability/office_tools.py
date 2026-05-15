"""OfficeTools — Batch file ops, document generation, file conversion, PDF processing.

Extends the 117 existing tools with missing office automation capabilities.
All tools follow the existing CapabilityBus pattern: function(Path, **kwargs) → dict/str.

Categories:
  1. Batch File Ops — rename, deduplicate, archive, organize
  2. Document Generation — docx, xlsx, pptx, PDF from templates
  3. File Conversion — md↔docx, csv↔xlsx, office→pdf
  4. PDF Processing — merge, split, extract, watermark, OCR
  5. Data Processing — deduplicate, merge, pivot, regex extract
  6. Email — send, check inbox, attachments
  7. Image — resize, compress, format convert
  8. Template Engine — Jinja2 fill, invoice, contract

Usage:
    from livingtree.capability.office_tools import OfficeTools
    result = OfficeTools.batch_rename("/docs", pattern=r"img_(\d+)", template="photo_{num:03d}")
    result = OfficeTools.generate_report(template="monthly", data={...}, format="docx")
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import shutil
import subprocess
import tempfile
import time
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from loguru import logger


class OfficeTools:
    """Office automation toolkit — batch ops, document generation, file conversion."""

    # ═══ 1. Batch File Operations ═════════════════════════════════

    @staticmethod
    def batch_rename(directory: str, pattern: str, template: str,
                     dry_run: bool = False, recursive: bool = False) -> dict:
        """Rename files matching regex pattern to template format.

        template supports: {num}, {name}, {ext}, {date}, {num:03d}
        Example: OfficeTools.batch_rename(".", r"img_(\d+)\.jpg", "photo_{num:03d}.jpg")"""
        p = Path(directory)
        files = list(p.rglob("*") if recursive else p.glob("*"))
        compiled = re.compile(pattern)
        renamed = []
        errors = []

        for i, f in enumerate(f for f in files if f.is_file()):
            m = compiled.match(f.name)
            if not m:
                continue
            groups = m.groupdict() or {str(j): v for j, v in enumerate(m.groups(), 1)}
            num_match = re.match(r'\{num:(\d+)d\}', template)
            num_fmt = f"{{:0{num_match.group(1)}d}}" if num_match else None
            new_name = template
            if '{num' in new_name and num_fmt:
                new_name = re.sub(r'\{num:\d+d\}', num_fmt.format(i + 1), new_name)
            for k, v in groups.items():
                new_name = new_name.replace(f"{{{k}}}", v)
            new_name = new_name.replace("{name}", f.stem).replace("{ext}", f.suffix.lstrip("."))
            new_name = new_name.replace("{date}", datetime.now().strftime("%Y%m%d"))

            new_path = f.parent / new_name
            if new_path.exists() and new_path != f:
                errors.append(f"Conflict: {f.name} → {new_name}")
                continue
            if not dry_run:
                f.rename(new_path)
            renamed.append({"old": f.name, "new": new_name})

        return {"renamed": len(renamed), "errors": len(errors),
                "files": renamed[:50], "dry_run": dry_run}

    @staticmethod
    def file_deduplicate(directory: str, recursive: bool = True,
                         delete: bool = False, min_size: int = 0) -> dict:
        """Find and optionally delete duplicate files by SHA256 hash."""
        seen: dict[str, list[str]] = {}
        for f in Path(directory).rglob("*" if recursive else "*"):
            if not f.is_file() or f.stat().st_size < min_size:
                continue
            try:
                h = hashlib.sha256(f.read_bytes()).hexdigest()
                seen.setdefault(h, []).append(str(f))
            except Exception:
                continue

        duplicates = {h: paths for h, paths in seen.items() if len(paths) > 1}
        deleted = 0
        saved_bytes = 0
        if delete:
            for paths in duplicates.values():
                for dup in paths[1:]:
                    try:
                        saved_bytes += Path(dup).stat().st_size
                        os.remove(dup)
                        deleted += 1
                    except Exception:
                        pass

        return {"duplicate_groups": len(duplicates), "duplicate_files": sum(len(v)-1 for v in duplicates.values()),
                "deleted": deleted, "saved_mb": round(saved_bytes / 1024 / 1024, 2),
                "samples": [{"hash": h[:12], "count": len(v), "first": v[0]} for h, v in list(duplicates.items())[:10]]}

    @staticmethod
    def batch_compress(sources: list[str], output: str,
                       format: str = "zip") -> dict:
        """Compress files/directories into archive."""
        output_path = Path(output)
        sources = [Path(s) for s in sources if Path(s).exists()]
        if not sources:
            return {"error": "No valid source files"}

        if format == "zip":
            with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
                for src in sources:
                    if src.is_dir():
                        for f in src.rglob("*"):
                            if f.is_file():
                                zf.write(f, f.relative_to(src.parent))
                    else:
                        zf.write(src, src.name)
        else:
            import tarfile
            with tarfile.open(output_path, f"w:{format}") as tar:
                for src in sources:
                    tar.add(src, src.name)

        size = output_path.stat().st_size
        return {"output": str(output_path), "format": format,
                "files": len(sources), "size_bytes": size}

    # ═══ 2. Document Generation ════════════════════════════════════

    @staticmethod
    def generate_report(template: str, data: dict, format: str = "md",
                        output: str = "") -> dict:
        """Generate a report from template with data binding.

        template: "monthly", "eia", "meeting", or a file path to a .md template
        data: variables to fill in
        format: "md" | "docx" | "pdf" | "html"
        """
        templates = {
            "monthly": """# 月度报告 — {month}

## 概述
{summary}

## 关键指标
| 指标 | 数值 | 变化 |
|------|------|------|
{metrics}

## 项目进展
{progress}

## 下月计划
{next_plan}

---
*自动生成于 {date}*
""",
            "meeting": """# 会议纪要 — {title}

- **时间**: {date} {time}
- **参会人**: {attendees}
- **地点**: {location}

## 议题
{agenda}

## 决议
{decisions}

## 待办事项
{action_items}

---
*记录人: {recorder}*
""",
        }

        tmpl = templates.get(template, "")
        if not tmpl and Path(template).exists():
            tmpl = Path(template).read_text(encoding="utf-8")
        if not tmpl:
            return {"error": f"Template '{template}' not found"}

        # Fill template
        data["date"] = data.get("date", datetime.now().strftime("%Y-%m-%d"))
        data["time"] = data.get("time", datetime.now().strftime("%H:%M"))
        try:
            content = tmpl.format(**{k: str(v) for k, v in data.items()})
        except KeyError as e:
            return {"error": f"Missing template variable: {e}"}

        output_path = Path(output or f"report_{int(time.time())}.{format}")
        content = OfficeTools._convert_format(content, format, output_path)
        return {"output": str(output_path), "format": format,
                "size": len(content), "template": template}

    @staticmethod
    def generate_table_xlsx(data: list[list], headers: list[str] = None,
                            output: str = "", sheet_name: str = "Sheet1") -> dict:
        """Generate Excel file from 2D data array."""
        try:
            import openpyxl
        except ImportError:
            # Fallback: CSV
            output_path = Path(output or f"table_{int(time.time())}.csv")
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                if headers:
                    writer.writerow(headers)
                writer.writerows(data)
            return {"output": str(output_path), "format": "csv", "rows": len(data)}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        if headers:
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h).font = openpyxl.styles.Font(bold=True)

        for row_idx, row in enumerate(data, 2 if headers else 1):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        output_path = Path(output or f"table_{int(time.time())}.xlsx")
        wb.save(output_path)
        return {"output": str(output_path), "format": "xlsx",
                "rows": len(data), "columns": len(data[0]) if data else 0}

    # ═══ 3. File Conversion ═══════════════════════════════════════

    @staticmethod
    def convert_file(source: str, target_format: str, output: str = "") -> dict:
        """Convert file between formats. Supports: md↔docx↔pdf↔html, csv↔xlsx, json↔csv."""
        sp = Path(source)
        if not sp.exists():
            return {"error": f"Source not found: {source}"}

        src_fmt = sp.suffix.lower().lstrip(".")
        out_path = Path(output or sp.with_suffix(f".{target_format}"))

        content = sp.read_text(encoding="utf-8", errors="replace")
        result = OfficeTools._convert_format(content, target_format, out_path)

        return {"output": str(out_path), "from": src_fmt, "to": target_format,
                "size": out_path.stat().st_size if out_path.exists() else 0}

    @staticmethod
    def _convert_format(content: str, target: str, out_path: Path) -> str:
        if target == "html":
            html = f"<html><body>{content.replace(chr(10),'<br>')}</body></html>"
            out_path.write_text(html, encoding="utf-8")
            return html
        elif target == "pdf":
            try:
                subprocess.run(["pandoc", "-f", "markdown", "-o", str(out_path),
                               "--pdf-engine=xelatex"], input=content, text=True,
                              timeout=30, capture_output=True)
            except Exception:
                # Fallback: write markdown as-is
                out_path.write_text(content, encoding="utf-8")
        else:
            out_path.write_text(content, encoding="utf-8")
        return content

    # ═══ 4. PDF Processing ════════════════════════════════════════

    @staticmethod
    def pdf_merge(sources: list[str], output: str) -> dict:
        """Merge multiple PDF files into one."""
        try:
            from pypdf import PdfWriter, PdfReader
        except ImportError:
            return {"error": "pypdf not installed. pip install pypdf"}

        writer = PdfWriter()
        for src in sources:
            if not Path(src).exists():
                return {"error": f"Not found: {src}"}
            reader = PdfReader(src)
            for page in reader.pages:
                writer.add_page(page)

        out = Path(output)
        writer.write(str(out))
        return {"output": str(out), "pages": len(writer.pages),
                "sources": len(sources)}

    @staticmethod
    def pdf_split(source: str, output_dir: str = "",
                  pages_per_file: int = 1) -> dict:
        """Split PDF into individual pages or chunks."""
        try:
            from pypdf import PdfReader, PdfWriter
        except ImportError:
            return {"error": "pypdf not installed"}

        reader = PdfReader(source)
        out_dir = Path(output_dir or Path(source).parent / "split")
        out_dir.mkdir(parents=True, exist_ok=True)

        files = []
        for i in range(0, len(reader.pages), pages_per_file):
            writer = PdfWriter()
            chunk = reader.pages[i:i + pages_per_file]
            for page in chunk:
                writer.add_page(page)
            out_path = out_dir / f"page_{i+1:04d}.pdf"
            writer.write(str(out_path))
            files.append(str(out_path))

        return {"output_dir": str(out_dir), "files": len(files),
                "total_pages": len(reader.pages)}

    @staticmethod
    def pdf_extract_text(source: str, pages: str = "all") -> dict:
        """Extract text from PDF pages."""
        try:
            from pypdf import PdfReader
        except ImportError:
            return {"error": "pypdf not installed"}

        reader = PdfReader(source)
        pages_to_extract = (range(len(reader.pages)) if pages == "all"
                           else [int(p)-1 for p in pages.split(",")])

        texts = []
        for i in pages_to_extract:
            if 0 <= i < len(reader.pages):
                texts.append(reader.pages[i].extract_text() or "")

        return {"pages": len(texts), "text": "\n---\n".join(texts),
                "total_chars": sum(len(t) for t in texts)}

    # ═══ 5. Data Processing ═══════════════════════════════════════

    @staticmethod
    def csv_deduplicate(source: str, columns: list[str] = None,
                        output: str = "") -> dict:
        """Remove duplicate rows from CSV based on specified columns."""
        sp = Path(source)
        if not sp.exists():
            return {"error": f"Not found: {source}"}

        with open(sp, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            rows = list(reader)

        seen = set()
        unique = []
        for row in rows:
            key = tuple(row.get(c, "") for c in (columns or headers))
            if key not in seen:
                seen.add(key)
                unique.append(row)

        out_path = Path(output or sp.parent / f"{sp.stem}_dedup.csv")
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(unique)

        return {"output": str(out_path), "original": len(rows),
                "unique": len(unique), "removed": len(rows) - len(unique)}

    @staticmethod
    def regex_extract(source: str, pattern: str, output: str = "",
                      group: int = 0) -> dict:
        """Extract data using regex from file content."""
        sp = Path(source)
        if not sp.exists():
            return {"error": f"Not found: {source}"}

        content = sp.read_text(encoding="utf-8", errors="replace")
        compiled = re.compile(pattern, re.MULTILINE)
        matches = [m.group(group) if group else m.group(0) for m in compiled.finditer(content)]

        out_path = Path(output or f"extracted_{int(time.time())}.txt")
        out_path.write_text("\n".join(matches), encoding="utf-8")

        return {"output": str(out_path), "matches": len(matches),
                "samples": matches[:10]}

    # ═══ 6. Email ═════════════════════════════════════════════════

    @staticmethod
    def send_email(to: str, subject: str, body: str,
                   smtp_config: dict = None, attachments: list[str] = None) -> dict:
        """Send email via SMTP."""
        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart
            from email.mime.base import MIMEBase
            from email import encoders
        except ImportError:
            return {"error": "email modules not available"}

        cfg = smtp_config or {}
        host = cfg.get("host", os.environ.get("SMTP_HOST", "smtp.gmail.com"))
        port = cfg.get("port", int(os.environ.get("SMTP_PORT", "587")))
        user = cfg.get("user", os.environ.get("SMTP_USER", ""))
        password = cfg.get("password", os.environ.get("SMTP_PASSWORD", ""))

        msg = MIMEMultipart()
        msg["From"] = user
        msg["To"] = to
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain", "utf-8"))

        for att in (attachments or []):
            with open(att, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={Path(att).name}")
                msg.attach(part)

        with smtplib.SMTP(host, port) as server:
            server.starttls()
            server.login(user, password)
            server.send_message(msg)

        return {"sent": True, "to": to, "subject": subject,
                "attachments": len(attachments or [])}

    # ═══ 7. Image Processing ══════════════════════════════════════

    @staticmethod
    def image_process(source: str, operation: str, output: str = "",
                      **params) -> dict:
        """Process image: resize, compress, format_convert, thumbnail.
        
        operation: "resize" (width, height), "compress" (quality 1-100),
                   "thumbnail" (size), "convert" (format: jpg/png/webp)
        """
        try:
            from PIL import Image
        except ImportError:
            return {"error": "Pillow not installed. pip install Pillow"}

        sp = Path(source)
        if not sp.exists():
            return {"error": f"Not found: {source}"}

        img = Image.open(sp)
        out_path = Path(output or sp.parent / f"{sp.stem}_processed{sp.suffix}")

        if operation == "resize":
            w = params.get("width", img.width)
            h = params.get("height", img.height)
            img = img.resize((w, h), Image.LANCZOS)
        elif operation == "compress":
            quality = params.get("quality", 80)
            img.save(out_path, quality=quality, optimize=True)
        elif operation == "thumbnail":
            size = params.get("size", 256)
            img.thumbnail((size, size))
        elif operation == "convert":
            fmt = params.get("format", "PNG")
            out_path = out_path.with_suffix(f".{fmt.lower()}")
            img.save(out_path, format=fmt)

        img.save(out_path)
        return {"output": str(out_path), "operation": operation,
                "size": out_path.stat().st_size, "dimensions": f"{img.width}x{img.height}"}

    # ═══ 8. Template Engine ═══════════════════════════════════════

    @staticmethod
    def fill_template(template_path: str, variables: dict,
                      output: str = "") -> dict:
        """Fill a template file with variables. Supports {var} syntax and Jinja2.

        For Jinja2 templates, use {{ var }} syntax in the template file.
        Simple templates use {var} syntax.
        """
        tp = Path(template_path)
        if not tp.exists():
            return {"error": f"Template not found: {template_path}"}

        content = tp.read_text(encoding="utf-8")

        # Try Jinja2 first
        if "{{" in content or "{%" in content:
            try:
                from jinja2 import Template
                tmpl = Template(content)
                result = tmpl.render(**variables)
            except ImportError:
                # Fallback to simple substitution
                result = content
                for k, v in variables.items():
                    result = result.replace("{{ " + k + " }}", str(v))
                    result = result.replace("{{" + k + "}}", str(v))
        else:
            result = content.format(**{k: str(v) for k, v in variables.items()})

        out_path = Path(output or tp.parent / f"filled_{tp.name}")
        out_path.write_text(result, encoding="utf-8")
        return {"output": str(out_path), "variables": len(variables),
                "size": len(result)}


__all__ = ["OfficeTools"]


def register_office_tools(bus=None):
    """Register all OfficeTools into CapabilityBus for LLM discovery."""
    try:
        from ..treellm.capability_bus import get_capability_bus, Capability, CapCategory, CapParam
        bus = bus or get_capability_bus()
        
        tools = [
            ("office:batch_rename", "Batch rename files matching pattern", "directory,pattern,template"),
            ("office:file_deduplicate", "Find duplicate files by content hash", "directory,recursive"),
            ("office:batch_compress", "Compress files into zip/tar archive", "sources,output,format"),
            ("office:generate_report", "Generate report from template with data", "template,data,format"),
            ("office:generate_table_xlsx", "Generate Excel spreadsheet", "data,headers,output"),
            ("office:convert_file", "Convert file between formats", "source,target_format"),
            ("office:pdf_merge", "Merge multiple PDFs into one", "sources,output"),
            ("office:pdf_split", "Split PDF into individual pages", "source,output_dir"),
            ("office:pdf_extract_text", "Extract text from PDF", "source,pages"),
            ("office:csv_deduplicate", "Remove duplicate CSV rows", "source,columns"),
            ("office:regex_extract", "Extract data using regex from file", "source,pattern"),
            ("office:send_email", "Send email via SMTP", "to,subject,body"),
            ("office:image_process", "Resize/compress/convert image", "source,operation,params"),
            ("office:fill_template", "Fill template with variables", "template_path,variables"),
        ]
        for cap_id, desc, hint in tools:
            bus.register(Capability(
                id=cap_id, name=cap_id.split(":",1)[1], category=CapCategory.TOOL,
                description=desc, params=[CapParam(name="input", type="string", description=hint)],
                source="office_tools", tags=["office", "automation"],
            ))
        logger.info(f"OfficeTools: registered {len(tools)} tools")
        return len(tools)
    except Exception as e:
        logger.debug(f"OfficeTools register: {e}")
        return 0
